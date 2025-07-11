import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import datetime

def cargar_inventario():
    try:
        wb = openpyxl.load_workbook('inventario_tecnologia.xlsx')
    except FileNotFoundError:
        print("¡Error! No se encontró el archivo de inventario.")
        return None, None  # Retornar dos Nones en lugar de solo uno.
    
    ws = wb.active
    ws.title = "Inventario Principal"

    #Aplicar Formato a los encabezados
    for col in range(1, ws.max_column + 1):
        celda = ws.cell(row=1, column=col)
        celda.font = Font(bold=True)                                            #color letra en negrita
        celda.fill = PatternFill(start_color="CCCCCC", fill_type="solid")       #color letra gris y con fondo gris
        celda.alignment = Alignment(horizontal="center")                        #alineación letra centrada

    return wb, ws

def actualizar_precio(ws, porcentaje):
    """ACtualiza los precios del inventario según un porcentaje"""
    for row in range(2, ws.max_row +1):
        precio_actual = float(ws.cell(row, column=4).value)
        nuevo_precio = precio_actual * (1 + porcentaje /100)
        ws.cell(row=row, column=4).value = round(nuevo_precio, 2)

def verificar_stock_bajo(ws, limite_minimo=10):
    """Generar alertas para productos con stock bajo"""
    productos_bajos = []
    for row in range(2, ws.max_row + 1):
        try:
            # Verificar y convertir el valor de stock a un número entero
            stock = int(ws.cell(row=row, column=5).value)
            if stock <= limite_minimo:
                producto = ws.cell(row=row, column=2).value
                productos_bajos.append((producto, stock))

            #Marcar en rojo las filas con stock bajo
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="FFB6B6",
                    fill_type="solid"
                )
        except (ValueError, TypeError):
            # Si el valor de stock no es un número válido, lo ignoramos
            continue

    return productos_bajos     

def generar_reporte():
    """Genera un reporte del inventario actual"""
    wb, ws = cargar_inventario()
    if not wb:
        return
    #Crear una hoja nueva para el reporte
    fecha_actual = datetime.datetime.now().strftime("%Y-%m-%d")
    ws_reporte = wb.create_sheet(f"Reporte {fecha_actual}")

    #Copiar encabezado
    for col in range(1, ws.max_column + 1):
        ws_reporte.cell(row=1, column=col).value = ws.cell(row=1, column=col).value
        ws_reporte.cell(row=1, column=col).font = Font(bold=True) 
    
    #copiar Datos
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws_reporte.cell(row=row, column=col).value = ws.cell(row=row, column=col).value

    #Añadir estadistica 
    row_stats = ws.max_row + 2
    ws_reporte.cell(row=row_stats, column=1, value="Estadistica del invenatior").font = Font(bold=True)
    ws_reporte.cell(row=row_stats + 1, column=1, value="Productos Diferentes")
    ws_reporte.cell(row=row_stats + 1, column=2, value=f"=COUNTA(B2:B{ws.max_row})")
    ws_reporte.cell(row=row_stats + 2, column=1, value="Total del inventario:")
    ws_reporte.cell(row=row_stats + 2, column=2, value=f"=SUM(E2:E{ws.max_row})")
    ws_reporte.cell(row=row_stats + 3, column=1, value=" Valor Total del inventario:")
    ws_reporte.cell(row=row_stats + 3, column=2, value=f"=SUMPRODUCT(D2:D{ws.max_row}, E2:E{ws.max_row})")

    return wb

def automatizacion_inventario():
    """Función principal que ejecuta las operaciones de automatización"""
    wb, ws = cargar_inventario()
    if not wb:
        return
    print("\n=== Sistema de Automatización de Inventario ===")

    #actualizar precios (ejemplo: aumento 5%)
    actualizar_precio(ws, porcentaje=5)
    print("\nPrecios actualizados con un incremento del 5%")

    #Verificar stock bajo
    productos_bajos = verificar_stock_bajo(ws)
    if productos_bajos:
        print("\n¡ALERTA! Productos con Stock Bajo:")
        for producto, stock in productos_bajos:
            print(f"- {producto}: {stock} unidades")

    wb = generar_reporte()
    
    wb.save('Inventario_tecnologia.xlsx')
    print("\nProceso de automatización completado exitosamente")

if __name__ == "__main__":
    automatizacion_inventario()