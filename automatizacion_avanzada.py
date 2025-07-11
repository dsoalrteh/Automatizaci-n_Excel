import openpyxl
import openpyxl.chart
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from collections import defaultdict
from openpyxl.chart.label import DataLabelList
import os

def crear_grafico_ventas(ws):
    """Crear gráfico de barras para visualizar stock por producto"""
    if ws.max_row < 2:
        print("No hay suficientes datos para crear el gráfico.")
        return


    productos = []
    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        producto = row[1]  # PRODUCTO
        stock = row[4]     # STOCK
        if producto and isinstance(stock, (int, float)):
            productos.append((producto, stock))                                #producto, stock

    #ordenar de Mayor a menor
    productos.sort(key=lambda x: x[1], reverse=True)

    # Crear una hoja temporal (no visible) para graficar ordenadamente
    temp_ws = ws.parent["TEMP_GRAFICO_PRODUCTO"] if "TEMP_GRAFICO_PRODUCTO" in ws.parent.sheetnames else ws.parent.create_sheet("TEMP_GRAFICO_PRODUCTO")
    temp_ws.delete_rows(1, temp_ws.max_row)

    temp_ws.append(["Producto", "Stock Total"])
    for prod, stock in productos:
        temp_ws.append([prod, stock])

    #crear grafico de barras
    grafico = BarChart()
    grafico.title = "Stock por producto"
    grafico.y_axis.title = 'Cantidad en Stock'
    grafico.x_axis.title = 'Productos'
    grafico.width = 40
    grafico.height = 15

    #Datos para el gráfico
    datos = Reference(temp_ws, min_col=2, min_row=1, max_row=len(productos)+1)
    categorias = Reference(temp_ws, min_col=1, min_row=2, max_row=len(productos)+1)

    grafico.add_data(datos, titles_from_data=True)
    grafico.set_categories(categorias)

    # Añadir etiquetas de datos
    for serie in grafico.series:
        serie.dLbls = openpyxl.chart.label.DataLabelList()
        serie.dLbls.showVal = True                                                          # Mostrar el valor de cada barra

    #Añadir gráfico a la hoja
    ws.add_chart(grafico, "G2")


def crear_grafico_categorias(ws):

    # Contar cantidad de productos por categoría
    conteo_stock = defaultdict(int)
    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        categoria = row[2]
        stock = row[4]
        if categoria and isinstance(stock, (int, float)):
            conteo_stock[categoria] += stock

        if not conteo_stock:
            print("No hay datos suficientes para el gráfico de categorías.")
            return
        
    # 2. Escribir los datos en hoja auxiliar temporal
    wb = ws.parent
    if "TEMP_GRAFICO" not in wb.sheetnames:
        temp_ws = wb.create_sheet("TEMP_GRAFICO")
    else: 
        temp_ws = wb["TEMP_GRAFICO"]
        temp_ws.delete_rows(1, temp_ws.max_row)

    temp_ws.append(["Categoria", "Stock Total"])
    for cat, total in conteo_stock.items():
        temp_ws.append([cat, total])    

    """Crear gráfico circular para visualizar distribución por categotías"""
    #crear grafico circular
    grafico = PieChart()
    grafico.title = "Disctribución por Categorías"

    #Datos para el gráfico
    datos = Reference(temp_ws, min_col=2, min_row=1, max_row=1 + len(conteo_stock))
    etiquetas = Reference(temp_ws, min_col=1, min_row=2, max_row=1 + len(conteo_stock))

    grafico.add_data(datos, titles_from_data=True)
    grafico.set_categories(etiquetas)

    # Añadir etiquetas de datos
    grafico.dataLabels = DataLabelList()
    grafico.dataLabels.showVal = True                                                       # Mostrar valor de cada sector

    #Añadir gráfico a la hoja
    ws.add_chart(grafico, "G33")

def crear_table(wb, ws):
    """Crear tabla"""
    #convertir rango de datos en tabla
    tab = Table(displayName="TablaInventario", ref=f"A1:E{ws.max_row}")

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )

    tab.tableStyleInfo = style
    ws.add_table(tab)

def aplicar_validacion_datos(ws):
    """Aplicar validación de datos a las columnas"""
    #Validación por categorías
    categorias_val = DataValidation(
        type="list",
        formula1='"Computadoras,Smartphones,Accesorios,Monitores,Perifericos"',
        allow_blank=False
    )
    categorias_val.error = "Por favor selecciones una categoría válida"
    categorias_val.errorTitle = "Categoría Invalida"
    ws.add_data_validation(categorias_val)
    categorias_val.add(f"C2:C{ws.max_row}")

    #Validación para stock (números positivos)
    stock_val = DataValidation(
        type="whole",
        operator="greaterThan",
        formula1="0"
    )
    stock_val.error = "El stock debe ser número positivo"
    stock_val.errorTitle = "Stock Invalido"
    ws.add_data_validation(stock_val)
    stock_val.add(f"E2:E{ws.max_row}")

def aplicar_formatos_condicionales(ws):
    """Aplicar formatos condicionales avanzados"""
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

    #Escala de Colores para precios
    color_scale = ColorScaleRule(
        start_type='min', start_color='09EE90',                 #color verde claro
        end_type='max', end_color='FF6B6B'                      #color rojo claro
    )

    ws.conditional_formatting.add(f"D2:D{ws.max_row}", color_scale)

def automatizacion_avanzada():
    """Función principal para la automatización avanzada"""
    #Cargar el archivo existente
    wb = openpyxl.load_workbook('inventario_tecnologia.xlsx')
    ws = wb.active

    # Verificar los nombres de las hojas
    print("Nombres de hojas en el archivo:", wb.sheetnames)

    #aplicar todas las funciones avanzadas
    print("1. Creando gráficos automaticos...")
    crear_grafico_ventas(ws)
    crear_grafico_categorias(ws)

    #crear tabla
    print("2. Creando Tabla...")
    crear_table(wb, ws)

    #Aplicando validación de datos en excel
    print("3. Aplicando validación de datos...")
    aplicar_validacion_datos(ws)

    #Aplicar formato condicional
    print("4. Aplicando formatos condicionales...")
    aplicar_formatos_condicionales(ws)

 
        # Guardar cambios
    try:
        wb.save('inventario_tecnologia.xlsx')
        print("Archivo guardado exitosamente.")
    except Exception as e:
        print(f"Error al guardar el archivo: {e}")

    #Si no deseas mantener la hoja TEMP_GRAFICO
    if "TEMP_GRAFICO" in wb.sheetnames:
        del wb["TEMP_GRAFICO"]

if __name__ == "__main__":
    automatizacion_avanzada()