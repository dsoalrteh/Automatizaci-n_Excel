import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def crear_excel_inicial():
    #Crear in nuevo libro de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Principal"

    #Definir los encabezados
    encabezados = ["id", "producto", "categoria", "precio", "stock"]

    datos = [
        [1001, "Computadora Dell", "Computadoras", 800.00, 50],
        [1002, "iPhone 13", "Smartphones", 999.99, 30],
        [1003, "Monitor Samsung", "Monitores", 300.00, 20],
        [1004, "Teclado Logitech", "Perifericos", 50.00, 100],
        [1005, "Mouse Razer", "Perifericos", 70.00, 75],
        [1006, "Cargador Anker", "Accesorios", 25.00, 150],
        [1007, "Batería Externa Aukey", "Accesorios", 40.00, 120],
        [1008, "Alfombrilla de ratón", "Perifericos", 10.00, 200],
        [1009, "Webcam Logitech", "Accesorios", 90.00, 50],
        [1010, "Computadora HP", "Computadoras", 700.00, 60],
        [1011, "Samsung Galaxy S21", "Smartphones", 899.00, 40],
        [1012, "Monitor LG", "Monitores", 250.00, 30],
        [1013, "Teclado Corsair", "Perifericos", 120.00, 80],
        [1014, "Mouse Logitech", "Perifericos", 30.00, 110],
        [1015, "Cargador Aukey", "Accesorios", 20.00, 160],
        [1016, "Batería Externa Belkin", "Accesorios", 45.00, 140],
        [1017, "Alfombrilla SteelSeries", "Perifericos", 15.00, 180],
        [1018, "Webcam Microsoft", "Accesorios", 65.00, 95],
        [1019, "Computadora Acer", "Computadoras", 750.00, 55],
        [1020, "iPhone 12", "Smartphones", 799.99, 35],
        [1021, "Monitor Asus", "Monitores", 280.00, 25],
        [1022, "Teclado Razer", "Perifericos", 100.00, 70],
        [1023, "Mouse SteelSeries", "Perifericos", 60.00, 85],
        [1024, "Cargador Belkin", "Accesorios", 18.00, 140],
        [1025, "Batería Externa Anker", "Accesorios", 50.00, 130],
        [1026, "Alfombrilla Logitech", "Perifericos", 20.00, 160],
        [1027, "Webcam Logitech Pro", "Accesorios", 120.00, 45],
        [1028, "Computadora Lenovo", "Computadoras", 650.00, 70],
        [1029, "Samsung Galaxy Note 20", "Smartphones", 1099.99, 25],
        [1030, "Monitor BenQ", "Monitores", 350.00, 40],
        [1031, "Teclado SteelSeries", "Perifericos", 95.00, 60],
        [1032, "Mouse Corsair", "Perifericos", 80.00, 90],
        [1033, "Cargador Samsung", "Accesorios", 22.00, 150],
        [1034, "Batería Externa Belkin", "Accesorios", 35.00, 110],
        [1035, "Alfombrilla Corsair", "Perifericos", 18.00, 170],
        [1036, "Webcam Logitech C920", "Accesorios", 100.00, 50],
        [1037, "Computadora MSI", "Computadoras", 950.00, 40],
        [1038, "iPhone SE", "Smartphones", 399.00, 60],
        [1039, "Monitor Acer", "Monitores", 200.00, 65],
        [1040, "Teclado Razer BlackWidow", "Perifericos", 150.00, 30],
        [1041, "Mouse Logitech G502", "Perifericos", 50.00, 110],
        [1042, "Cargador iOttie", "Accesorios", 35.00, 100],
        [1043, "Batería Externa PowerCore", "Accesorios", 60.00, 100],
        [1044, "Alfombrilla Logitech G640", "Perifericos", 12.00, 180],
        [1045, "Webcam Logitech Brio", "Accesorios", 150.00, 40],
        [1046, "Computadora Dell Inspiron", "Computadoras", 700.00, 50],
        [1047, "Samsung Galaxy Z Fold 3", "Smartphones", 1799.99, 15],
        [1048, "Monitor Dell", "Monitores", 320.00, 50],
        [1049, "Teclado HyperX Alloy", "Perifericos", 110.00, 45],
        [1050, "Mouse Razer DeathAdder", "Perifericos", 90.00, 75],
        [1051, "Cargador Mophie", "Accesorios", 35.00, 100],
        [1052, "Batería Externa Zendure", "Accesorios", 55.00, 120],
        [1053, "Alfombrilla Roccat", "Perifericos", 25.00, 150],
        [1054, "Webcam Logitech StreamCam", "Accesorios", 160.00, 30],
        [1055, "Computadora ASUS ROG", "Computadoras", 1200.00, 20],
        [1056, "iPhone 11", "Smartphones", 749.00, 50],
        [1057, "Monitor ViewSonic", "Monitores", 270.00, 40],
        [1058, "Teclado Logitech G Pro", "Perifericos", 130.00, 70],
        [1059, "Mouse HyperX Pulsefire", "Perifericos", 60.00, 90],
        [1060, "Cargador RavPower", "Accesorios", 20.00, 140],
        [1061, "Batería Externa Goal Zero", "Accesorios", 65.00, 80],
        [1062, "Alfombrilla SteelSeries QcK", "Perifericos", 15.00, 160],
        [1063, "Webcam Creative Labs", "Accesorios", 110.00, 50],
        [1064, "Computadora Gigabyte", "Computadoras", 800.00, 60],
        [1065, "iPhone XR", "Smartphones", 649.00, 55],
        [1066, "Monitor Eizo", "Monitores", 650.00, 30],
        [1067, "Teclado Razer Huntsman", "Perifericos", 180.00, 25],
        [1068, "Mouse Logitech G203", "Perifericos", 30.00, 120],
        [1069, "Cargador UGREEN", "Accesorios", 25.00, 130],
        [1070, "Batería Externa Romoss", "Accesorios", 40.00, 110],
        [1071, "Alfombrilla Corsair MM300", "Perifericos", 14.00, 160],
        [1072, "Webcam Logitech C922", "Accesorios", 95.00, 60],
        [1073, "Computadora Microsoft", "Computadoras", 850.00, 50],
        [1074, "iPhone 10", "Smartphones", 600.00, 70],
        [1075, "Monitor Philips", "Monitores", 250.00, 40],
        [1076, "Teclado Redragon", "Perifericos", 45.00, 100]
    ]

    #Escribir encabezados
    for col, encabezado in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col)
        celda.value = encabezado
        celda.font = Font(bold=True)
        celda.fill = PatternFill(start_color="CCCCCC", fill_type="solid")
        celda.alignment = Alignment(horizontal="center")

    #Escribir datos
    for row_idx, fila in enumerate(datos, 2):
        for col_idx, valor in enumerate(fila, 1):
            celda = ws.cell(row=row_idx, column=col_idx)
            celda.value = valor
            celda.fill = PatternFill(start_color="CCCCCC", fill_type="solid")

    #Ajustar ancho de columnas
    for col in range(1, len(encabezados) +1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    #Guardar el archivo
    wb.save('inventario_tecnologia.xlsx')
    print("Archivo Excel creado exitosamente")

if __name__ == '__main__':
    crear_excel_inicial()